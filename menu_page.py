import flet as ft
from openpyxl import load_workbook
import threading
import os
import trader_type1 as trader1
from flet import (
    Column,
    ElevatedButton,
    Page,
    Row,
    TextField,
    UserControl,
    Dropdown,
    icons,
    dropdown,
)
from enum import Enum


class Options(Enum):
    One = 1
    All = 2


class Menu(UserControl):
    def __init__(self, page: Page):
        super().__init__()
        self.page = page

    def build(self):
        self.new_task = TextField(
            hint_text="Привет",
            expand=True,
        )
        self.products = Dropdown()
        self.trader_type = Dropdown(
            on_change=self.dropdown_changed,
            options=[
                ft.dropdown.Option(Options(1).name),
            ],
            width=200,
        )
        return Column(
            [
                ft.Text("Menu"),
                Row(
                    [
                        self.products,
                        self.trader_type,
                        ElevatedButton(
                            "Start trade",
                            icon=icons.PHOTO_SIZE_SELECT_ACTUAL,
                            width=160,
                            on_click=self.change_trader,
                        ),
                    ]
                ),
            ],
            height=500,
            alignment=ft.MainAxisAlignment.START,
            expand=True,
            spacing=20,
        )

    def dropdown_changed(self, e):
        super().update()

    def change_trader(self, e):
        try:
            if self.trader_type.value == Options(1).name:
                parts = self.products.value.split(":")
                my_thread = threading.Thread(
                    target=lambda: trader1.main(
                        productName=parts[0], productPrice=int(parts[1])
                    )
                )
                my_thread.start()
                my_thread.join()

        except Exception as e:
            print(f"Ошибка чтения файла: {e}")

    def did_mount(self):
        file_name = "products.xlsx"
        file_path = os.path.join(
            os.path.dirname(__file__), file_name
        )  # Определение пути к файлу относительно текущего скрипта
        try:
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                filtered_row = [str(cell) for cell in row if cell is not None]
                if filtered_row:
                    combined_row = ":".join(filtered_row)
                    self.products.options.append(dropdown.Option(combined_row))
            super().update()
        except Exception as e:
            print(f"Ошибка чтения файла: {e}")
