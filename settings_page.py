import flet as ft
import threading
import screen_selector as sc
import os
import json
import trader_type1 as trader_type1
import openpyxl
from flet import (
    Column,
    ElevatedButton,
    Page,
    Row,
    UserControl,
    icons,
)


import screen_selector as sc


class Settings(UserControl):
    def __init__(self, page: Page):
        super().__init__()
        self.page = page

    def build(self):
        self.field_path = ft.TextField(
            label="Путь до файла tesseract.exe",
            value=self.page.client_storage.get("tesseract_path"),
            width=500,
        )
        self.x = ft.TextField(
            label="x",
            value=self.page.client_storage.get("x"),
            disabled=True,
            width=100,
        )

        self.y = ft.TextField(
            label="y",
            value=self.page.client_storage.get("y"),
            disabled=True,
            width=100,
        )

        self.width = ft.TextField(
            label="width",
            value=self.page.client_storage.get("width"),
            disabled=True,
            width=100,
        )

        self.height = ft.TextField(
            label="height",
            value=self.page.client_storage.get("height"),
            disabled=True,
            width=100,
        )
        self.product_name = ft.TextField(
            text_size=30,
            cursor_color=ft.colors.RED,
            selection_color=ft.colors.YELLOW,
            color=ft.colors.PINK,
            bgcolor=ft.colors.BLACK26,
            filled=True,
            focused_color=ft.colors.GREEN,
            focused_bgcolor=ft.colors.CYAN_200,
            border_radius=30,
            border_color=ft.colors.GREEN_800,
            focused_border_color=ft.colors.GREEN_ACCENT_400,
            max_length=20,
            capitalization="characters",
        )
        self.product_price = ft.TextField(
            text_size=30,
            cursor_color=ft.colors.RED,
            selection_color=ft.colors.YELLOW,
            color=ft.colors.PINK,
            bgcolor=ft.colors.BLACK26,
            filled=True,
            focused_color=ft.colors.GREEN,
            focused_bgcolor=ft.colors.CYAN_200,
            border_radius=30,
            border_color=ft.colors.GREEN_800,
            focused_border_color=ft.colors.GREEN_ACCENT_400,
            max_length=20,
            capitalization="characters",
        )

        return Column(
            [
                ft.Text("Settings"),
                Row(
                    [
                        ElevatedButton(
                            "Path to tesseract",
                            icon=icons.PHOTO_SIZE_SELECT_ACTUAL,
                            width=160,
                            on_click=self.pick_files_result,
                        ),
                        self.field_path,
                    ],
                    spacing=10,
                ),
                ft.Divider(height=1),
                Row(
                    [
                        ElevatedButton(
                            "Pick Box Size",
                            icon=icons.PHOTO_SIZE_SELECT_ACTUAL,
                            width=160,
                            on_click=self.screen_select,
                        ),
                        self.x,
                        self.y,
                        self.width,
                        self.height,
                    ],
                    spacing=10,
                ),
                ft.Divider(height=1),
                Row(
                    [
                        ElevatedButton(
                            "Add product",
                            icon=icons.PHOTO_SIZE_SELECT_ACTUAL,
                            width=160,
                            on_click=self.add_product_to_excel,
                        ),
                        self.product_name,
                        self.product_price,
                    ],
                    spacing=10,
                ),
            ],
            height=500,
            alignment=ft.MainAxisAlignment.START,
            expand=True,
            spacing=20,
        )

    def screen_select(self, e):
        my_thread = threading.Thread(target=sc.main())
        my_thread.start()
        my_thread.join()

        current_directory = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_directory, "ar_data.json")

        with open(file_path, "r") as json_file:
            data = json.load(json_file)
            self.page.client_storage.set("x", data["x"])
            self.page.client_storage.set("y", data["y"])
            self.page.client_storage.set("width", data["width"])
            self.page.client_storage.set("height", data["height"])
            self.x.value = data["x"]
            self.y.value = data["y"]
            self.width.value = data["width"]
            self.height.value = data["height"]
            self.x.update()
            self.y.update()
            self.width.update()
            self.height.update()

    def pick_files_result(self, e):
        selected_file_path = self.field_path.value
        if os.path.exists(selected_file_path):
            self.page.client_storage.set("tesseract_path", selected_file_path)
            super().update()
        else:
            raise Exception("Path does not exist")

    def add_product_to_excel(self, e):
        try:
            file_name = "products.xlsx"
            file_path = os.path.join(os.path.dirname(__file__), file_name)

            if not self.product_name.value and self.product_price.value:
                raise ValueError("Некорректные данные: отсутствует название продукта.")

            if not os.path.exists(file_path):
                wb = openpyxl.Workbook()
                sheet = wb.active
            else:
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

            next_row = sheet.max_row + 1

            sheet.cell(row=next_row, column=1).value = self.product_name.value
            sheet.cell(row=next_row, column=2).value = self.product_price.value
            self.product_name.value = None
            self.product_price.value = None
            super().update()
            wb.save(file_path)
            print("Данные успешно добавлены в файл Excel.")
        except Exception as e:
            print(f"Ошибка при добавлении данных в файл Excel: {e}")
